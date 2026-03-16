from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from algorithms.utils.visualization import plot_dict_curves
import pandas as pd
from collections import defaultdict
import numpy as np
from algorithms.evaluation.fitness_evaluator import FitnessEvaluator
import config
import os
from datetime import datetime
import pandas as pd
import numpy as np



def log_all_individuals_binary(islands, dataset):
    """Loggt ALLE aktuellen Individuen pro Insel."""
    feature_cols = dataset.df_preprocessed.columns.tolist()
    all_data = []

    generation = max([max(list(island.best_fitness_per_gen.keys())) for island in islands])
    timestamp = datetime.now().strftime("%H:%M:%S")

    for island_id, island in enumerate(islands):
        current_pop = island.population  # ← DEAP Population
        for rank, ind in enumerate(current_pop, 1):
            binary_str = ','.join(map(str, ind))
            all_data.append({
                'timestamp': timestamp,
                'generation': generation,
                'island_id': island_id,
                'ind_rank': rank,
                'fitness': ind.fitness.values[0],
                'num_features': sum(ind),
                'binary_vector': binary_str
            })

    df_all = pd.DataFrame(all_data)
    df_all.to_csv("logs/all_individuals_binary.csv", index=False)


def process_results(islands,results_folder, results_file_path, dataset, sa_island):
    # Required: results_df, settings, evaltime, fitness, avg_diversity, migration_info
    write_gen_info(islands, dataset.target, results_folder, results_file_path)
    write_results_full(results_file_path, islands, dataset)

    if sa_island is not None:
        write_sa_log(sa_island, results_file_path, dataset.target)
        log_all_individuals_binary(islands, dataset)

def get_min_fitness_per_gen(islands):
    gen_to_fitnesses = defaultdict(list)
    for island in islands:
        for gen, fitness in island.best_fitness_per_gen.items():
            gen_to_fitnesses[gen].append(fitness)

    # Ermittle den besten (kleinsten) Fitnesswert je Generation
    best_fitness_per_gen = []
    for gen in sorted(gen_to_fitnesses.keys()):
        best_fitness = min(gen_to_fitnesses[gen])
        best_fitness_per_gen.append(best_fitness)

    return best_fitness_per_gen

def get_avg_diverstiy(islands):
    gen_to_diverstiy = defaultdict(list)
    for island in islands:
        for gen, diversity in island.population_diversity_per_gen.items():
            gen_to_diverstiy[gen].append(diversity)

    avg_diversity_per_gen = []
    for gen in sorted(gen_to_diverstiy.keys()):
        avg_diversity = sum(gen_to_diverstiy[gen]) / len(gen_to_diverstiy[gen])
        avg_diversity_per_gen.append(avg_diversity)

    return avg_diversity_per_gen

def write_gen_info(islands, target, results_folder, results_file_path):
    sheet_name = f"Stats_{target}"
    best_fitness_per_gen = get_min_fitness_per_gen(islands)
    diversity_per_gen = get_avg_diverstiy(islands)

    generations = list(range(1, len(best_fitness_per_gen) + 1))

    # Erstelle den DataFrame
    df_results = pd.DataFrame({
        'Gen': generations,
        'Best Fitness': best_fitness_per_gen,
        'Avg Diversity': diversity_per_gen
    })

    settings = {
        'Target': target,
        'Rule Type': config.rule_type,
        'Group Type': config.group_type,
        'Number of GA Islands': config.num_islands,
        'Total Starting Pop Size': config.pop_size,
        'Migration Interval': config.migration_interval,
        'Iterations': config.iterations,
        'Start Population Strategy': config.start_population_strategy,
        'Topology': config.topology,
        'Migration Strategy': config.migration_strategy,
    }
    df_settings = pd.DataFrame(list(settings.items()), columns=["Setting", "Value"])

    best_fitness_png = plot_best_fitness(islands, target, results_folder)
    avg_diversity_png = plot_avg_diversity(islands, target, results_folder)
    population_size_png = plot_population_size(islands, target, results_folder)

    with pd.ExcelWriter(results_file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df_results.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
        df_settings.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=4)

        worksheet = writer.book[sheet_name]

        img1 = Image(best_fitness_png)
        img1.width = img1.width * 0.3
        img1.height = img1.height * 0.3
        worksheet.add_image(img1, 'H2')

        img2 = Image(avg_diversity_png)
        img2.width = img2.width * 0.3
        img2.height = img2.height * 0.3
        worksheet.add_image(img2, 'H25')

        img3 = Image(population_size_png)
        img3.width = img3.width * 0.3
        img3.height = img3.height * 0.3
        worksheet.add_image(img3, 'H48')

def plot_best_fitness(islands, target, results_folder):
    best_fitness_data = []
    for island in islands:
        best_fitness_data.append(island.best_fitness_per_gen)

    file_path = plot_dict_curves(best_fitness_data, title="Best Fitness Over Generations", ylabel="Fitness", results_folder=results_folder, target=target)
    return file_path

def plot_avg_diversity(islands, target, results_folder):
    diversity_data = []
    for island in islands:
        diversity_data.append(island.population_diversity_per_gen)

    file_path = plot_dict_curves(diversity_data, title="Average Diversity Over Generations", ylabel="Diversity", results_folder=results_folder, target=target)
    return file_path

def plot_population_size(islands, target, results_folder):
    population_size_data = []
    for island in islands:
        population_size_data.append(island.population_size)

    file_path = plot_dict_curves(population_size_data, title="Population Size Over Generations", ylabel="Population Size", results_folder=results_folder, target=target)
    return file_path

def write_results_full(results_file_path, islands, dataset):
    results_df = create_results_df(islands, dataset)

    sheet_name = f"full_result_{dataset.target}"
    wb = load_workbook(results_file_path)

    if sheet_name in wb.sheetnames:
        # If sheet exists, append data to the existing sheet
        with pd.ExcelWriter(results_file_path, engine='openpyxl', mode='a') as writer:
            writer = pd.ExcelWriter(results_file_path, engine='openpyxl', mode='a', if_sheet_exists="overwrite")
            results_df.to_excel(writer, sheet_name=sheet_name, index=False, header=None, startrow=wb[sheet_name].max_row)
    else:
        # Else create a new sheet and write data
        with pd.ExcelWriter(results_file_path, engine='openpyxl', mode='a') as writer:
            results_df.to_excel(writer, sheet_name=sheet_name, index=False)

def write_sa_log(sa_island, results_file_path, target):
    sa_log_df = pd.DataFrame(sa_island.sa_log)

    sheet_name = f"SA_Log_{target}"
    wb = load_workbook(results_file_path)

    if sheet_name in wb.sheetnames:
        # If sheet exists, append data to the existing sheet
        with pd.ExcelWriter(results_file_path, engine='openpyxl', mode='a') as writer:
            writer.book = wb
            sa_log_df.to_excel(writer, sheet_name=sheet_name, index=False, header=None, startrow=wb[sheet_name].max_row)
    else:
        # Else create a new sheet and write data
        with pd.ExcelWriter(results_file_path, engine='openpyxl', mode='a') as writer:
            sa_log_df.to_excel(writer, sheet_name=sheet_name, index=False)

def create_results_df(islands, dataset):
    best_individuals = get_top_x_individuals(islands, 10)

    best_individuals_array = [np.append(individual[:-1], [0]) for individual in best_individuals]

    results_df = pd.DataFrame(data=best_individuals_array, columns=dataset.df_preprocessed.columns)

    sum_best_ind = [np.sum(individual[:-1]) for individual in best_individuals_array]
    results_df.insert(0, "IND_SUM", sum_best_ind, allow_duplicates=True)

    # Add column containing the fitness scores with penalty for to many activated features of selected individuals
    individual_scores = [ind.fitness.values[0] for ind in best_individuals]
    results_df.insert(0, "EVAL_SCORE_PENALTY", individual_scores, allow_duplicates=True)

    # Add column containing the fitness scores without penalty for to many activated features of selected individuals
    evaluator = FitnessEvaluator(target=dataset.target, data_split=0.8, dataset=dataset.df_preprocessed, 
                                                max_features=dataset.max_features, penalty=True)
    individual_scores_no_penalty = [evaluator.evaluate(ind)[0] for ind in best_individuals_array]
    results_df.insert(0, "EVAL_SCORE_NO_PENALTY", individual_scores_no_penalty, allow_duplicates=True)

    # Add column with information about how often each feature was activated
    # new_row = pd.DataFrame(index=range(1), columns=results_df.columns)
    new_row = pd.DataFrame([{col: "" for col in results_df.columns}])
    results_df = pd.concat([results_df, new_row], ignore_index=True)
    results_df.reset_index(drop=True, inplace=True)
    # Preparing the data to print from info_dict
    sorted_keys = sorted(dataset.gen_dict)
    sublist = []
    for key in sorted_keys:
        value = dataset.gen_dict[key]
        sublist.append(value.memory)

    # add the selection count for each feature here
    for index, feature in enumerate(sublist):
        results_df.iloc[-1, 3 + index] = str(feature[dataset.target])

    # Add column referring the solutions to the target estimated
    results_df.insert(0, "REG_PARAMETER", dataset.target, allow_duplicates=True)

    return results_df

def get_top_x_individuals(islands, x):
    """
    Get the top x individuals from all islands. No duplicates are allowed.
    NOTE: Islands are expected to have a hall of fame with the best individuals, therefore should be GAIslands.
    """
    all_individuals = []
    for island in islands:
        all_individuals.extend(island.halloffame[:x])
    
    # Dont allow duplicates
    unique_individuals = list({tuple(ind): ind for ind in all_individuals}.values())

    # Sort by fitness
    unique_individuals.sort(key=lambda ind: ind.fitness.values[0])
    
    return unique_individuals[:x]